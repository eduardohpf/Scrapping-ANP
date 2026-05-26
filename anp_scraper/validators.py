"""Validação do arquivo exportado."""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from anp_scraper.exceptions import ValidationError

log = logging.getLogger(__name__)

MIN_XLSX_BYTES = 1024
LARGE_FILE_BYTES = 5 * 1024 * 1024
SAMPLE_MAX_ROWS = 100


def _row_has_data(row: tuple) -> bool:
    return any(cell is not None and str(cell).strip() != "" for cell in row)


def validate_export_file(path: Path) -> dict:
    """
    Valida que o download foi concluído e contém planilha legível.

    Returns:
        dict com metadados (tamanho, linhas, colunas).
    """
    if not path.exists():
        raise ValidationError(f"Arquivo não encontrado: {path}")

    size = path.stat().st_size
    if size < MIN_XLSX_BYTES:
        raise ValidationError(f"Arquivo muito pequeno ({size} bytes): {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise ValidationError(f"Extensão inesperada {suffix!r} em {path}")

    try:
        read_only = size > LARGE_FILE_BYTES
        wb = load_workbook(path, read_only=read_only, data_only=True)
        ws = wb.active
        row_count = 0
        col_count = 0
        scanned = 0

        for row in ws.iter_rows(values_only=True):
            scanned += 1
            if _row_has_data(row):
                row_count += 1
                col_count = max(col_count, len(row))
            if read_only and row_count >= 2 and scanned >= SAMPLE_MAX_ROWS:
                break

        wb.close()
    except Exception as exc:
        raise ValidationError(f"Não foi possível ler Excel: {path}") from exc

    if row_count < 2:
        raise ValidationError(f"Planilha sem dados suficientes (linhas={row_count})")

    meta: dict = {
        "path": str(path),
        "size_bytes": size,
        "rows": row_count if not read_only else None,
        "columns": col_count,
    }
    if read_only:
        meta["rows_note"] = f"amostra_{SAMPLE_MAX_ROWS}_linhas_arquivo_grande"

    log.info(
        "Exportação validada: %s (%s bytes, linhas=%s, colunas=%s)",
        path.name,
        size,
        meta["rows"] if meta["rows"] is not None else "n/a(grande)",
        col_count,
    )
    return meta
